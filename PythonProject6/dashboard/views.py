# dashboard/views.py
from django.shortcuts import render, redirect
from django.db import connection
from .forms import LoginForm
import hashlib
import datetime


def login(request):
    """Страница входа"""
    error = None

    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # ИСПРАВЛЕНО: Используем правильную кодировку UTF-8
            password_hash = hashlib.sha256(password.encode('utf-8')).hexdigest()

            print(f"🔍 Хеш SHA256 (UTF-8) от '{password}': {password_hash}")

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT user_id, username, full_name, role_id 
                    FROM Users 
                    WHERE username = %s AND password_hash = %s AND is_active = 1
                """, [username, password_hash])
                user = cursor.fetchone()

                if user:
                    request.session['user_id'] = user[0]
                    request.session['username'] = user[1]
                    request.session['full_name'] = user[2]
                    request.session['role_id'] = user[3]
                    request.session['is_authenticated'] = True

                    print(f"✅ Успешный вход! Пользователь: {user[2]}")
                    return redirect('dashboard:index')
                else:
                    error = 'Неверный логин или пароль'
                    print("❌ Пользователь не найден в базе данных")
                    # Дополнительная отладка
                    print(f"🔍 Проверка в базе: пользователь '{username}' с хешем '{password_hash}'")
                    with connection.cursor() as cursor2:
                        cursor2.execute("SELECT COUNT(*) FROM Users WHERE username = %s", [username])
                        user_exists = cursor2.fetchone()[0]
                        print(f"🔍 Пользователь '{username}' существует в базе: {user_exists > 0}")
                        if user_exists > 0:
                            cursor2.execute("SELECT password_hash FROM Users WHERE username = %s", [username])
                            stored_hash = cursor2.fetchone()[0]
                            print(f"🔍 Хеш в базе для '{username}': {stored_hash}")
                            print(f"🔍 Совпадают ли хеши: {password_hash == stored_hash}")
    else:
        form = LoginForm()

    return render(request, 'dashboard/login.html', {'form': form, 'error': error})


def logout(request):
    """Выход из системы"""
    request.session.flush()
    return redirect('dashboard:login')


# ==================== СТРАНИЦА ПАРТИЙ ====================
def batches(request):
    """Страница управления партиями"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    # Получаем параметры фильтрации
    status_filter = request.GET.get('status', '')
    supplier_filter = request.GET.get('supplier', '')
    search_query = request.GET.get('search', '')

    with connection.cursor() as cursor:
        # Получаем все партии с фильтрацией
        query = """
            SELECT 
                pb.batch_id,
                p.name as product_name,
                s.name as supplier_name,
                pb.batch_number,
                pb.production_date,
                pb.expiration_date,
                pb.quantity,
                bs.status_name,
                pb.received_at
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            JOIN Suppliers s ON pb.supplier_id = s.supplier_id
            JOIN BatchStatuses bs ON pb.batch_status_id = bs.batch_status_id
            WHERE 1=1
        """
        params = []

        if status_filter:
            query += " AND bs.status_name = %s"
            params.append(status_filter)

        if supplier_filter:
            query += " AND s.name = %s"
            params.append(supplier_filter)

        if search_query:
            query += " AND (p.name LIKE %s OR s.name LIKE %s)"
            params.extend([f'%{search_query}%', f'%{search_query}%'])

        query += " ORDER BY pb.received_at DESC"

        cursor.execute(query, params)
        batches = cursor.fetchall()

        # Получаем уникальных поставщиков для фильтра
        cursor.execute("SELECT DISTINCT name FROM Suppliers ORDER BY name")
        suppliers = [row[0] for row in cursor.fetchall()]

        # Получаем статусы для фильтра
        cursor.execute("SELECT DISTINCT status_name FROM BatchStatuses")
        statuses = [row[0] for row in cursor.fetchall()]

    context = {
        'batches': batches,
        'suppliers': suppliers,
        'statuses': statuses,
        'status_filter': status_filter,
        'supplier_filter': supplier_filter,
        'search_query': search_query,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/batches.html', context)


def add_batch(request):
    """Добавление новой партии"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        supplier_id = request.POST.get('supplier_id')
        batch_number = request.POST.get('batch_number')
        production_date = request.POST.get('production_date')
        expiration_date = request.POST.get('expiration_date')
        quantity = request.POST.get('quantity')
        mode_id = request.POST.get('mode_id')

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO ProductBatches 
                (product_id, supplier_id, batch_number, production_date, expiration_date, quantity, batch_status_id, mode_id)
                VALUES (%s, %s, %s, %s, %s, %s, 1, %s)
            """, [product_id, supplier_id, batch_number, production_date, expiration_date, quantity, mode_id])

        return redirect('dashboard:batches')

    with connection.cursor() as cursor:
        # Исправленный запрос: используем правильное название столбца 'name'
        cursor.execute("SELECT mode_id, name FROM temperaturemodes ORDER BY name")
        temperature_modes = cursor.fetchall()

        # Отладочный вывод
        print("Таблица temperaturemodes:", temperature_modes)

        cursor.execute("SELECT product_id, name FROM Products WHERE is_active = 1 ORDER BY name")
        products = cursor.fetchall()

        cursor.execute("SELECT supplier_id, name FROM Suppliers ORDER BY name")
        suppliers = cursor.fetchall()

    # Отладочный вывод в контекст
    context = {
        'products': products,
        'suppliers': suppliers,
        'temperature_modes': temperature_modes,
        'debug': f"Таблица temperaturemodes: {temperature_modes}",
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/add_batch.html', context)


def edit_batch(request, batch_id):
    """Редактирование партии"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    with connection.cursor() as cursor:
        if request.method == 'POST':
            batch_number = request.POST.get('batch_number')
            production_date = request.POST.get('production_date')
            expiration_date = request.POST.get('expiration_date')
            quantity = request.POST.get('quantity')
            mode_id = request.POST.get('mode_id')

            cursor.execute("""
                UPDATE ProductBatches 
                SET batch_number = %s, production_date = %s, expiration_date = %s, quantity = %s, mode_id = %s
                WHERE batch_id = %s
            """, [batch_number, production_date, expiration_date, quantity, mode_id, batch_id])

            return redirect('dashboard:batches')

        cursor.execute("""
            SELECT 
                pb.batch_id,
                pb.product_id,
                pb.supplier_id,
                pb.batch_number,
                pb.production_date,
                pb.expiration_date,
                pb.quantity,
                pb.mode_id,
                p.name as product_name,
                s.name as supplier_name
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            JOIN Suppliers s ON pb.supplier_id = s.supplier_id
            WHERE pb.batch_id = %s
        """, [batch_id])
        batch = cursor.fetchone()

        cursor.execute("SELECT product_id, name FROM Products WHERE is_active = 1 ORDER BY name")
        products = cursor.fetchall()

        cursor.execute("SELECT supplier_id, name FROM Suppliers ORDER BY name")
        suppliers = cursor.fetchall()

        # Исправленный запрос: используем правильное название столбца 'name'
        cursor.execute("SELECT mode_id, name FROM temperaturemodes ORDER BY name")
        temperature_modes = cursor.fetchall()

    context = {
        'batch': batch,
        'products': products,
        'suppliers': suppliers,
        'temperature_modes': temperature_modes,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/edit_batch.html', context)


def delete_batch(request, batch_id):
    """Удаление партии"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM ProductBatches WHERE batch_id = %s", [batch_id])

    return redirect('dashboard:batches')


# ==================== СТРАНИЦА ПРОДУКТОВ ====================
def products(request):
    """Страница управления продуктами"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    # Получаем параметры фильтрации
    category_filter = request.GET.get('category', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    search_query = request.GET.get('search', '')

    with connection.cursor() as cursor:
        # Получаем все продукты с фильтрацией
        query = """
            SELECT 
                p.product_id,
                p.name,
                p.sku,
                p.price,
                c.name as category_name,
                p.description,
                p.is_active
            FROM Products p
            LEFT JOIN Categories c ON p.category_id = c.category_id
            WHERE p.is_active = 1
        """
        params = []

        if category_filter:
            query += " AND c.name = %s"
            params.append(category_filter)

        if min_price:
            query += " AND p.price >= %s"
            params.append(min_price)

        if max_price:
            query += " AND p.price <= %s"
            params.append(max_price)

        if search_query:
            query += " AND p.name LIKE %s"
            params.append(f'%{search_query}%')

        query += " ORDER BY p.name"

        cursor.execute(query, params)
        products = cursor.fetchall()

        # Получаем категории для фильтра
        cursor.execute("SELECT DISTINCT name FROM Categories ORDER BY name")
        categories = [row[0] for row in cursor.fetchall()]

    context = {
        'products': products,
        'categories': categories,
        'category_filter': category_filter,
        'min_price': min_price,
        'max_price': max_price,
        'search_query': search_query,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/products.html', context)


def add_product(request):
    """Добавление нового продукта"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    if request.method == 'POST':
        name = request.POST.get('name')
        sku = request.POST.get('sku')
        price = request.POST.get('price')
        category_id = request.POST.get('category_id')
        description = request.POST.get('description')
        mode_id = request.POST.get('mode_id')  # ← ДОБАВЛЕНО

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO Products (name, sku, price, category_id, description, mode_id, is_active)
                VALUES (%s, %s, %s, %s, %s, %s, 1)
            """, [name, sku, price, category_id, description, mode_id])

        return redirect('dashboard:products')

    with connection.cursor() as cursor:
        cursor.execute("SELECT category_id, name FROM Categories ORDER BY name")
        categories = cursor.fetchall()

        # ← ДОБАВЛЕНО: получаем режимы хранения
        cursor.execute("SELECT mode_id, name FROM temperaturemodes ORDER BY name")
        temperature_modes = cursor.fetchall()

    context = {
        'categories': categories,
        'temperature_modes': temperature_modes,  # ← ДОБАВЛЕНО
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/add_product.html', context)


def edit_product(request, product_id):
    """Редактирование продукта"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    with connection.cursor() as cursor:
        if request.method == 'POST':
            name = request.POST.get('name')
            sku = request.POST.get('sku')
            price = request.POST.get('price')
            category_id = request.POST.get('category_id')
            description = request.POST.get('description')
            mode_id = request.POST.get('mode_id')  # ← ДОБАВЛЕНО

            cursor.execute("""
                UPDATE Products 
                SET name = %s, sku = %s, price = %s, category_id = %s, description = %s, mode_id = %s
                WHERE product_id = %s
            """, [name, sku, price, category_id, description, mode_id, product_id])

            return redirect('dashboard:products')

        # ← ДОБАВЛЕНО: получаем mode_id в запросе
        cursor.execute("""
            SELECT 
                p.product_id,
                p.name,
                p.sku,
                p.price,
                p.category_id,
                p.description,
                p.mode_id,  # ← ДОБАВЛЕНО
                c.name as category_name
            FROM Products p
            LEFT JOIN Categories c ON p.category_id = c.category_id
            WHERE p.product_id = %s
        """, [product_id])
        product = cursor.fetchone()

        cursor.execute("SELECT category_id, name FROM Categories ORDER BY name")
        categories = cursor.fetchall()

        # ← ДОБАВЛЕНО: получаем режимы хранения
        cursor.execute("SELECT mode_id, name FROM temperaturemodes ORDER BY name")
        temperature_modes = cursor.fetchall()

    context = {
        'product': product,
        'categories': categories,
        'temperature_modes': temperature_modes,  # ← ДОБАВЛЕНО
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/edit_product.html', context)


def delete_product(request, product_id):
    """Удаление продукта"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    with connection.cursor() as cursor:
        # Мягкое удаление (деактивация)
        cursor.execute("UPDATE Products SET is_active = 0 WHERE product_id = %s", [product_id])

    return redirect('dashboard:products')


# ==================== СИСТЕМА ОТЧЕТОВ ====================

def reports(request):
    """Главная страница отчетов"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    # Проверка прав доступа (только администратор и аналитик)
    role_id = request.session.get('role_id')
    if role_id not in [1, 5]:  # 1 = Администратор, 5 = Аналитик
        return redirect('dashboard:index')

    # Статистика для главной страницы отчетов
    with connection.cursor() as cursor:
        # Общая стоимость остатков
        cursor.execute("""
            SELECT COALESCE(SUM(pb.quantity * p.price), 0) 
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            WHERE pb.batch_status_id = 1
        """)
        total_stock_value = cursor.fetchone()[0]

        # Партии с истекающим сроком
        cursor.execute("""
            SELECT COUNT(*) FROM ProductBatches 
            WHERE expiration_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 3 DAY)
            AND batch_status_id = 1
        """)
        expiring_soon = cursor.fetchone()[0]

        # Общее количество списаний
        cursor.execute("SELECT COUNT(*) FROM WriteOffs")
        total_writeoffs = cursor.fetchone()[0]

        # Общая сумма списаний
        cursor.execute("""
            SELECT COALESCE(SUM(w.quantity * p.price), 0)
            FROM WriteOffs w
            JOIN ProductBatches pb ON w.batch_id = pb.batch_id
            JOIN Products p ON pb.product_id = p.product_id
        """)
        total_writeoff_value = cursor.fetchone()[0]

        # Топ-5 продуктов по остаткам
        cursor.execute("""
            SELECT p.name, SUM(pb.quantity) as total_qty, SUM(pb.quantity * p.price) as total_value
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            WHERE pb.batch_status_id = 1
            GROUP BY p.product_id, p.name
            ORDER BY total_value DESC
            LIMIT 5
        """)
        top_products = cursor.fetchall()

        # Топ-5 поставщиков по объему
        cursor.execute("""
            SELECT s.name, COUNT(pb.batch_id) as batch_count, SUM(pb.quantity) as total_qty
            FROM ProductBatches pb
            JOIN Suppliers s ON pb.supplier_id = s.supplier_id
            GROUP BY s.supplier_id, s.name
            ORDER BY total_qty DESC
            LIMIT 5
        """)
        top_suppliers = cursor.fetchall()

    context = {
        'total_stock_value': round(total_stock_value, 2),
        'expiring_soon': expiring_soon,
        'total_writeoffs': total_writeoffs,
        'total_writeoff_value': round(total_writeoff_value, 2),
        'top_products': top_products,
        'top_suppliers': top_suppliers,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/reports.html', context)


def report_stock(request):
    """Отчет по остаткам продукции"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    role_id = request.session.get('role_id')
    if role_id not in [1, 3, 5]:  # Администратор, Кладовщик, Аналитик
        return redirect('dashboard:reports')

    # Получаем параметры фильтрации
    category_filter = request.GET.get('category', '')
    min_qty = request.GET.get('min_qty', '')
    max_qty = request.GET.get('max_qty', '')

    with connection.cursor() as cursor:
        # Основной запрос с фильтрацией
        query = """
            SELECT 
                p.product_id,
                p.name as product_name,
                c.name as category_name,
                p.sku,
                p.price,
                SUM(pb.quantity) as total_quantity,
                SUM(pb.quantity * p.price) as total_value,
                COUNT(DISTINCT pb.batch_id) as batch_count
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            LEFT JOIN Categories c ON p.category_id = c.category_id
            WHERE pb.batch_status_id = 1
        """
        params = []

        if category_filter:
            query += " AND c.name = %s"
            params.append(category_filter)

        if min_qty:
            query += " AND pb.quantity >= %s"
            params.append(min_qty)

        if max_qty:
            query += " AND pb.quantity <= %s"
            params.append(max_qty)

        query += """
            GROUP BY p.product_id, p.name, c.name, p.sku, p.price
            ORDER BY total_value DESC
        """

        cursor.execute(query, params)
        stock_data = cursor.fetchall()

        # Получаем категории для фильтра
        cursor.execute("SELECT DISTINCT name FROM Categories ORDER BY name")
        categories = [row[0] for row in cursor.fetchall()]

        # Статистика по категориям
        cursor.execute("""
            SELECT 
                c.name,
                COUNT(DISTINCT p.product_id) as product_count,
                SUM(pb.quantity) as total_qty,
                SUM(pb.quantity * p.price) as total_value
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            JOIN Categories c ON p.category_id = c.category_id
            WHERE pb.batch_status_id = 1
            GROUP BY c.category_id, c.name
            ORDER BY total_value DESC
        """)
        category_stats = cursor.fetchall()

    context = {
        'stock_data': stock_data,
        'categories': categories,
        'category_filter': category_filter,
        'min_qty': min_qty,
        'max_qty': max_qty,
        'category_stats': category_stats,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/report_stock.html', context)


def report_writeoffs(request):
    """Отчет по списаниям"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    role_id = request.session.get('role_id')
    if role_id not in [1, 5]:  # Только Администратор и Аналитик
        return redirect('dashboard:reports')

    # Получаем параметры фильтрации
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    reason_filter = request.GET.get('reason', '')

    with connection.cursor() as cursor:
        # Основной запрос
        query = """
            SELECT 
                w.write_off_id,
                w.write_off_date,
                p.name as product_name,
                pb.batch_number,
                w.quantity,
                p.price,
                (w.quantity * p.price) as loss_amount,
                wor.reason_name,
                u.full_name as written_by,
                w.notes
            FROM WriteOffs w
            JOIN ProductBatches pb ON w.batch_id = pb.batch_id
            JOIN Products p ON pb.product_id = p.product_id
            JOIN WriteOffReasons wor ON w.write_off_reason_id = wor.write_off_reason_id
            JOIN Users u ON w.user_id = u.user_id
            WHERE 1=1
        """
        params = []

        if start_date:
            query += " AND DATE(w.write_off_date) >= %s"
            params.append(start_date)

        if end_date:
            query += " AND DATE(w.write_off_date) <= %s"
            params.append(end_date)

        if reason_filter:
            query += " AND wor.reason_name = %s"
            params.append(reason_filter)

        query += " ORDER BY w.write_off_date DESC"

        cursor.execute(query, params)
        writeoffs = cursor.fetchall()

        # Получаем причины списания для фильтра
        cursor.execute("SELECT DISTINCT reason_name FROM WriteOffReasons ORDER BY reason_name")
        reasons = [row[0] for row in cursor.fetchall()]

        # Статистика по причинам
        cursor.execute("""
            SELECT 
                wor.reason_name,
                COUNT(w.write_off_id) as count,
                SUM(w.quantity) as total_qty,
                SUM(w.quantity * p.price) as total_loss
            FROM WriteOffs w
            JOIN ProductBatches pb ON w.batch_id = pb.batch_id
            JOIN Products p ON pb.product_id = p.product_id
            JOIN WriteOffReasons wor ON w.write_off_reason_id = wor.write_off_reason_id
            WHERE 1=1
        """ + (" AND DATE(w.write_off_date) >= %s" if start_date else "") +
                       (" AND DATE(w.write_off_date) <= %s" if end_date else ""),
                       [x for x in [start_date, end_date] if x])

        reason_stats = cursor.fetchall()

        # Топ-5 продуктов по списаниям
        cursor.execute("""
            SELECT 
                p.name,
                COUNT(w.write_off_id) as writeoff_count,
                SUM(w.quantity) as total_qty,
                SUM(w.quantity * p.price) as total_loss
            FROM WriteOffs w
            JOIN ProductBatches pb ON w.batch_id = pb.batch_id
            JOIN Products p ON pb.product_id = p.product_id
            WHERE 1=1
        """ + (" AND DATE(w.write_off_date) >= %s" if start_date else "") +
                       (" AND DATE(w.write_off_date) <= %s" if end_date else "") +
                       """
                       GROUP BY p.product_id, p.name
                       ORDER BY total_loss DESC
                       LIMIT 5
                       """,
                       [x for x in [start_date, end_date] if x])

        top_products = cursor.fetchall()

    context = {
        'writeoffs': writeoffs,
        'reasons': reasons,
        'reason_filter': reason_filter,
        'start_date': start_date,
        'end_date': end_date,
        'reason_stats': reason_stats,
        'top_products': top_products,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/report_writeoffs.html', context)


def report_expirations(request):
    """Отчет по срокам годности"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    role_id = request.session.get('role_id')
    if role_id not in [1, 3, 5]:  # Администратор, Кладовщик, Аналитик
        return redirect('dashboard:reports')

    # Получаем параметры
    days_ahead = int(request.GET.get('days', 7))

    with connection.cursor() as cursor:
        # Партии с истекающим сроком
        cursor.execute("""
            SELECT 
                pb.batch_id,
                p.name as product_name,
                s.name as supplier_name,
                pb.batch_number,
                pb.production_date,
                pb.expiration_date,
                pb.quantity,
                p.price,
                (pb.quantity * p.price) as value,
                DATEDIFF(pb.expiration_date, CURDATE()) as days_left
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            JOIN Suppliers s ON pb.supplier_id = s.supplier_id
            WHERE pb.batch_status_id = 1
            AND pb.expiration_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL %s DAY)
            ORDER BY pb.expiration_date ASC
        """, [days_ahead])

        expiring_batches = cursor.fetchall()

        # Статистика по дням
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN DATEDIFF(pb.expiration_date, CURDATE()) <= 1 THEN 'Сегодня-завтра'
                    WHEN DATEDIFF(pb.expiration_date, CURDATE()) <= 3 THEN '2-3 дня'
                    WHEN DATEDIFF(pb.expiration_date, CURDATE()) <= 7 THEN '4-7 дней'
                    ELSE 'Более 7 дней'
                END as period,
                COUNT(pb.batch_id) as batch_count,
                SUM(pb.quantity) as total_qty,
                SUM(pb.quantity * p.price) as total_value
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            WHERE pb.batch_status_id = 1
            AND pb.expiration_date >= CURDATE()
            GROUP BY period
            ORDER BY 
                CASE period
                    WHEN 'Сегодня-завтра' THEN 1
                    WHEN '2-3 дня' THEN 2
                    WHEN '4-7 дней' THEN 3
                    ELSE 4
                END
        """)

        period_stats = cursor.fetchall()

        # Критические партии (истекает сегодня)
        cursor.execute("""
            SELECT 
                pb.batch_id,
                p.name as product_name,
                pb.batch_number,
                pb.expiration_date,
                pb.quantity,
                p.price,
                (pb.quantity * p.price) as value
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            WHERE pb.batch_status_id = 1
            AND DATE(pb.expiration_date) = CURDATE()
        """)

        critical_batches = cursor.fetchall()

    context = {
        'expiring_batches': expiring_batches,
        'period_stats': period_stats,
        'critical_batches': critical_batches,
        'days_ahead': days_ahead,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/report_expirations.html', context)


def report_sales(request):
    """Отчет по продажам"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    role_id = request.session.get('role_id')
    if role_id not in [1, 2, 5]:  # Администратор, Менеджер, Аналитик
        return redirect('dashboard:reports')

    # Получаем параметры
    period = request.GET.get('period', 'month')  # day, week, month, year

    with connection.cursor() as cursor:
        # Выручка по периодам
        if period == 'day':
            cursor.execute("""
                SELECT 
                    DATE(wo.created_at) as date,
                    COUNT(wo.order_id) as order_count,
                    SUM(wo.total_amount) as revenue,
                    AVG(wo.total_amount) as avg_check
                FROM WholesaleOrders wo
                WHERE wo.order_status_id IN (5, 6)  -- Доставлен, Отправлен
                AND DATE(wo.created_at) >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                GROUP BY DATE(wo.created_at)
                ORDER BY date DESC
            """)
        elif period == 'week':
            cursor.execute("""
                SELECT 
                    YEARWEEK(wo.created_at, 1) as week,
                    COUNT(wo.order_id) as order_count,
                    SUM(wo.total_amount) as revenue,
                    AVG(wo.total_amount) as avg_check
                FROM WholesaleOrders wo
                WHERE wo.order_status_id IN (5, 6)
                AND wo.created_at >= DATE_SUB(CURDATE(), INTERVAL 12 WEEK)
                GROUP BY YEARWEEK(wo.created_at, 1)
                ORDER BY week DESC
            """)
        elif period == 'month':
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(wo.created_at, '%%Y-%%m') as month,
                    COUNT(wo.order_id) as order_count,
                    SUM(wo.total_amount) as revenue,
                    AVG(wo.total_amount) as avg_check
                FROM WholesaleOrders wo
                WHERE wo.order_status_id IN (5, 6)
                AND wo.created_at >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(wo.created_at, '%%Y-%%m')
                ORDER BY month DESC
            """)
        else:  # year
            cursor.execute("""
                SELECT 
                    YEAR(wo.created_at) as year,
                    COUNT(wo.order_id) as order_count,
                    SUM(wo.total_amount) as revenue,
                    AVG(wo.total_amount) as avg_check
                FROM WholesaleOrders wo
                WHERE wo.order_status_id IN (5, 6)
                GROUP BY YEAR(wo.created_at)
                ORDER BY year DESC
            """)

        sales_data = cursor.fetchall()

        # Топ-10 продуктов по продажам
        cursor.execute("""
            SELECT 
                p.name,
                SUM(woi.quantity) as total_sold,
                SUM(woi.total_price) as total_revenue
            FROM WholesaleOrderItems woi
            JOIN Products p ON woi.product_id = p.product_id
            JOIN WholesaleOrders wo ON woi.order_id = wo.order_id
            WHERE wo.order_status_id IN (5, 6)
            GROUP BY p.product_id, p.name
            ORDER BY total_revenue DESC
            LIMIT 10
        """)

        top_products = cursor.fetchall()

        # Статистика по клиентам
        cursor.execute("""
            SELECT 
                wo.customer_name,
                COUNT(wo.order_id) as order_count,
                SUM(wo.total_amount) as total_spent,
                AVG(wo.total_amount) as avg_check
            FROM WholesaleOrders wo
            WHERE wo.order_status_id IN (5, 6)
            GROUP BY wo.customer_name
            ORDER BY total_spent DESC
            LIMIT 10
        """)

        top_customers = cursor.fetchall()

    context = {
        'sales_data': sales_data,
        'top_products': top_products,
        'top_customers': top_customers,
        'period': period,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/report_sales.html', context)


def report_suppliers(request):
    """Отчет по поставщикам"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    role_id = request.session.get('role_id')
    if role_id not in [1, 3, 5]:  # Администратор, Кладовщик, Аналитик
        return redirect('dashboard:reports')

    with connection.cursor() as cursor:
        # Статистика по поставщикам
        cursor.execute("""
            SELECT 
                s.supplier_id,
                s.name,
                s.contact_person,
                s.phone,
                s.email,
                COUNT(DISTINCT pb.batch_id) as batch_count,
                SUM(pb.quantity) as total_quantity,
                SUM(pb.quantity * p.price) as total_value,
                AVG(DATEDIFF(pb.expiration_date, pb.production_date)) as avg_shelf_life,
                (
                    SELECT COUNT(*) 
                    FROM WriteOffs w
                    JOIN ProductBatches pb2 ON w.batch_id = pb2.batch_id
                    WHERE pb2.supplier_id = s.supplier_id
                ) as writeoff_count
            FROM Suppliers s
            LEFT JOIN ProductBatches pb ON s.supplier_id = pb.supplier_id
            LEFT JOIN Products p ON pb.product_id = p.product_id
            WHERE s.is_active = 1
            GROUP BY s.supplier_id, s.name, s.contact_person, s.phone, s.email
            ORDER BY total_value DESC
        """)

        suppliers_data = cursor.fetchall()

        # Рейтинг поставщиков
        cursor.execute("""
            SELECT 
                s.name,
                COUNT(DISTINCT pb.batch_id) as batches,
                SUM(pb.quantity) as total_qty,
                (
                    SELECT COUNT(*) 
                    FROM WriteOffs w
                    JOIN ProductBatches pb2 ON w.batch_id = pb2.batch_id
                    WHERE pb2.supplier_id = s.supplier_id
                ) as writeoffs,
                CASE 
                    WHEN (
                        SELECT COUNT(*) 
                        FROM WriteOffs w
                        JOIN ProductBatches pb2 ON w.batch_id = pb2.batch_id
                        WHERE pb2.supplier_id = s.supplier_id
                    ) = 0 THEN 'Отлично'
                    WHEN (
                        SELECT COUNT(*) 
                        FROM WriteOffs w
                        JOIN ProductBatches pb2 ON w.batch_id = pb2.batch_id
                        WHERE pb2.supplier_id = s.supplier_id
                    ) < 3 THEN 'Хорошо'
                    ELSE 'Требует внимания'
                END as rating
            FROM Suppliers s
            LEFT JOIN ProductBatches pb ON s.supplier_id = pb.supplier_id
            WHERE s.is_active = 1
            GROUP BY s.supplier_id, s.name
            ORDER BY writeoffs ASC, total_qty DESC
        """)

        supplier_ratings = cursor.fetchall()

    context = {
        'suppliers_data': suppliers_data,
        'supplier_ratings': supplier_ratings,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/report_suppliers.html', context)


def report_temperature(request):
    """Отчет по температурному режиму"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    role_id = request.session.get('role_id')
    if role_id not in [1, 3, 5]:  # Администратор, Кладовщик, Аналитик
        return redirect('dashboard:reports')

    # Получаем параметры
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    with connection.cursor() as cursor:
        # Нарушения температурного режима
        query = """
            SELECT 
                tl.log_id,
                tl.timestamp,
                p.name as product_name,
                pb.batch_number,
                tl.temperature,
                tm.min_temp,
                tm.max_temp,
                ls.status_name,
                u.full_name as measured_by,
                tl.notes
            FROM TemperatureLogs tl
            JOIN ProductBatches pb ON tl.batch_id = pb.batch_id
            JOIN Products p ON pb.product_id = p.product_id
            JOIN TemperatureModes tm ON pb.mode_id = tm.mode_id
            JOIN LogStatuses ls ON tl.log_status_id = ls.log_status_id
            LEFT JOIN Users u ON tl.measured_by = u.user_id
            WHERE tl.log_status_id IN (2, 3)  -- warning, critical
        """
        params = []

        if start_date:
            query += " AND DATE(tl.timestamp) >= %s"
            params.append(start_date)

        if end_date:
            query += " AND DATE(tl.timestamp) <= %s"
            params.append(end_date)

        query += " ORDER BY tl.timestamp DESC"

        cursor.execute(query, params)
        violations = cursor.fetchall()

        # Статистика по статусам
        cursor.execute("""
            SELECT 
                ls.status_name,
                COUNT(tl.log_id) as count,
                AVG(tl.temperature) as avg_temp
            FROM TemperatureLogs tl
            JOIN LogStatuses ls ON tl.log_status_id = ls.log_status_id
            WHERE 1=1
        """ + (" AND DATE(tl.timestamp) >= %s" if start_date else "") +
                       (" AND DATE(tl.timestamp) <= %s" if end_date else ""),
                       [x for x in [start_date, end_date] if x])

        status_stats = cursor.fetchall()

        # Критические партии
        cursor.execute("""
            SELECT 
                pb.batch_id,
                p.name as product_name,
                pb.batch_number,
                pb.expiration_date,
                pb.quantity,
                COUNT(tl.log_id) as violation_count,
                MAX(tl.temperature) as max_temp,
                MIN(tl.temperature) as min_temp
            FROM TemperatureLogs tl
            JOIN ProductBatches pb ON tl.batch_id = pb.batch_id
            JOIN Products p ON pb.product_id = p.product_id
            WHERE tl.log_status_id = 3  -- critical
            """ + (" AND DATE(tl.timestamp) >= %s" if start_date else "") +
                       (" AND DATE(tl.timestamp) <= %s" if end_date else "") +
                       """
                       GROUP BY pb.batch_id, p.name, pb.batch_number, pb.expiration_date, pb.quantity
                       HAVING violation_count > 0
                       ORDER BY violation_count DESC
                       LIMIT 10
                   """,
                       [x for x in [start_date, end_date] if x])

        critical_batches = cursor.fetchall()

    context = {
        'violations': violations,
        'status_stats': status_stats,
        'critical_batches': critical_batches,
        'start_date': start_date,
        'end_date': end_date,
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/report_temperature.html', context)


def export_report(request, report_type):
    """Экспорт отчета в Excel"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    role_id = request.session.get('role_id')
    if role_id not in [1, 5]:  # Только Администратор и Аналитик
        return redirect('dashboard:reports')

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active

    # Заголовки в зависимости от типа отчета
    if report_type == 'stock':
        ws.title = "Остатки продукции"
        headers = ['ID', 'Продукт', 'Категория', 'Артикул', 'Цена', 'Количество', 'Стоимость', 'Партии']
        ws.append(headers)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    p.product_id,
                    p.name,
                    c.name,
                    p.sku,
                    p.price,
                    SUM(pb.quantity),
                    SUM(pb.quantity * p.price),
                    COUNT(DISTINCT pb.batch_id)
                FROM ProductBatches pb
                JOIN Products p ON pb.product_id = p.product_id
                LEFT JOIN Categories c ON p.category_id = c.category_id
                WHERE pb.batch_status_id = 1
                GROUP BY p.product_id, p.name, c.name, p.sku, p.price
                ORDER BY SUM(pb.quantity * p.price) DESC
            """)
            data = cursor.fetchall()

    elif report_type == 'writeoffs':
        ws.title = "Списания"
        headers = ['ID', 'Дата', 'Продукт', 'Партия', 'Количество', 'Цена', 'Сумма', 'Причина', 'Кто списал']
        ws.append(headers)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    w.write_off_id,
                    w.write_off_date,
                    p.name,
                    pb.batch_number,
                    w.quantity,
                    p.price,
                    (w.quantity * p.price),
                    wor.reason_name,
                    u.full_name
                FROM WriteOffs w
                JOIN ProductBatches pb ON w.batch_id = pb.batch_id
                JOIN Products p ON pb.product_id = p.product_id
                JOIN WriteOffReasons wor ON w.write_off_reason_id = wor.write_off_reason_id
                JOIN Users u ON w.user_id = u.user_id
                ORDER BY w.write_off_date DESC
            """)
            data = cursor.fetchall()

    elif report_type == 'expirations':
        ws.title = "Сроки годности"
        headers = ['ID', 'Продукт', 'Поставщик', 'Партия', 'Производство', 'Истечение', 'Количество', 'Цена',
                   'Стоимость', 'Дней осталось']
        ws.append(headers)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    pb.batch_id,
                    p.name,
                    s.name,
                    pb.batch_number,
                    pb.production_date,
                    pb.expiration_date,
                    pb.quantity,
                    p.price,
                    (pb.quantity * p.price),
                    DATEDIFF(pb.expiration_date, CURDATE())
                FROM ProductBatches pb
                JOIN Products p ON pb.product_id = p.product_id
                JOIN Suppliers s ON pb.supplier_id = s.supplier_id
                WHERE pb.batch_status_id = 1
                AND pb.expiration_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY)
                ORDER BY pb.expiration_date ASC
            """)
            data = cursor.fetchall()

    else:
        ws.title = "Отчет"
        headers = ['Данные']
        ws.append(headers)
        data = []

    # Стилизация заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Заполнение данными
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Автоширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Создание ответа
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename="report_{report_type}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response

def index(request):
    """Главная страница — доступна только авторизованным"""
    if not request.session.get('is_authenticated', False):
        return redirect('dashboard:login')

    with connection.cursor() as cursor:
        # Общее количество активных партий
        cursor.execute("""
            SELECT COUNT(*) FROM ProductBatches 
            WHERE batch_status_id = (
                SELECT batch_status_id FROM BatchStatuses WHERE status_name = 'active'
            )
        """)
        total_active_batches = cursor.fetchone()[0]

        # Общее количество продуктов
        cursor.execute("SELECT COUNT(*) FROM Products WHERE is_active = 1")
        total_products = cursor.fetchone()[0]

        # Партии с истекающим сроком (менее 3 дней)
        cursor.execute("""
            SELECT COUNT(*) FROM ProductBatches 
            WHERE expiration_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 3 DAY)
            AND batch_status_id = (
                SELECT batch_status_id FROM BatchStatuses WHERE status_name = 'active'
            )
        """)
        expiring_soon = cursor.fetchone()[0]

        # Общая стоимость остатков
        cursor.execute("""
            SELECT COALESCE(SUM(pb.quantity * p.price), 0) 
            FROM ProductBatches pb
            JOIN Products p ON pb.product_id = p.product_id
            WHERE pb.batch_status_id = (
                SELECT batch_status_id FROM BatchStatuses WHERE status_name = 'active'
            )
        """)
        total_stock_value = cursor.fetchone()[0]

    context = {
        'total_active_batches': total_active_batches,
        'total_products': total_products,
        'expiring_soon': expiring_soon,
        'total_stock_value': round(total_stock_value, 2),
        'username': request.session.get('full_name', request.session.get('username', 'Гость')),
    }

    return render(request, 'dashboard/index.html', context)